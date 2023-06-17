# data_deal.py 

from status import *

import openpyxl as pyxl
import pandas as pd
from math import *
import time

flog = open('log.txt','w')
start = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
def logging(s):
    s = str(s) # 转为字符串
    print(s)
    flog.write(s + '\n')

logging('')
logging(start)

""" list写法"""
class Dealer:
    def __init__(self, filename, app) -> None:
        # 打开现有工作簿
        # workbook = pyxl.load_workbook(filename='example.xlsx')
        # for sheet_name in workbook.sheetnames:
        #     sheet = workbook[sheet_name]
            
        self.filename_out = filename[:-5] + '-out' + '.xlsx'
        self.done = 0
        sheets_dict = pd.read_excel(filename, sheet_name=None)
        app.label1.config(text = "处理中...")
        app.root.update()
        # 先行创建一个新的excel文件
        workbook = pyxl.Workbook()
        workbook.save(self.filename_out)
        
        for sheet_name, df in sheets_dict.items():
            logging(f'\n工作表名称：{sheet_name}')
            if len(df.columns)<5:
                logging('not enough data')
                continue
            # 如果列数等于五，那么就在第一列加上一列顺排号
            if len(df.columns) == 5:
                pass
            if len(df.columns) > 6:
                df = df.iloc[:, :6] # 去除后面多余的列
            col_frame = df.columns[0]
            self.frames = df[col_frame].tolist()
            self.times = df[df.columns[1]].tolist()
            self.X1 = df[df.columns[2]].tolist()
            self.Y1 = df[df.columns[3]].tolist()
            self.X2 = df[df.columns[4]].tolist()
            self.Y2 = df.iloc[:, 5].tolist()
            self.df = df
            
            self.deal()
            self.sheetname = sheet_name
            self.output()
        pass
            
    def deal(self):
        self.K=[]
        self.angle=[]
        self.X_mid = []
        self.Y_mid=[]
        zerot = 0
        
        """计算 mid 和 angle"""
        for i in range(min(len(self.X1),len(self.X2))):
            if(self.X1[i]==0 or self.X2[i]==0):
                zerot += 1
                continue
            xmid = (self.X1[i]+self.X2[i])/2
            ymid = (self.Y1[i]+self.Y2[i])/2
            self.X_mid.append([i,xmid])
            self.Y_mid.append([i,ymid])
            if self.X2[i] - self.X1[i]==0:
                k=0
            else:
                k=(self.Y2[i]-self.Y1[i])/(self.X2[i]-self.X1[i])
            self.K.append(k)
            self.angle.append(atan(k)*180/pi)
            
            self.num = len(self.X_mid) # 最终处理的数量按照self.num来算
            
        """计算转向半径radius"""
        self.radius = []
        max_r = 0
        cnt1 = 0
        cnt2 = 0
        cnt3 = 0
        for i in range(len(self.X_mid) - 2):
            if self.X_mid[i+2][0]-self.X_mid[i+1][0] == 1 and self.X_mid[i+1][0]-self.X_mid[i][0] == 1: # 连续三点
                cnt1 += 1
                d_s = sqrt((self.Y_mid[i+1][1]-self.Y_mid[i][1])**2 +(self.X_mid[i+1][1]-self.X_mid[i][1])**2)
                d_thres = 1e-5
                if d_s > d_thres: # 两点间距离差大于阈值
                    cnt2 += 1
                    alpha1 = (atan((self.Y_mid[i+2][1] - self.Y_mid[i+1][1]) / (self.X_mid[i+2][1] - self.X_mid[i+1][1])) 
                            if abs(self.X_mid[i+2][1] - self.X_mid[i+1][1]) > d_thres else pi/2)
                    alpha2 = (atan((self.Y_mid[i+1][1] - self.Y_mid[i][1]) / (self.X_mid[i+1][1] - self.X_mid[i][1])) 
                            if abs(self.X_mid[i+1][1] - self.X_mid[i][1]) > d_thres else pi/2)
                    d_alpha = alpha1 - alpha2
                    if d_alpha > 1e-3:
                        cnt3 += 1
                        r = d_s / d_alpha
                        """changed: withdrew;"""
                        self.radius.append(r)
                        if r > max_r:
                            max_r = r
                    else:
                        self.radius.append(0)
                else:
                    self.radius.append(0)
            else: # these three situation we can't calculate the radius, default 0
                self.radius.append(0)
        assert len(self.radius) == self.num - 2, ValueError("Wrong Value")
        logging(f'filter: {self.num}, {cnt1}, {cnt2}, {cnt3}')

        
    def output(self):
        angle_df = pd.DataFrame(self.angle, columns=['转向角度'])
        radius_df = pd.DataFrame(self.radius, columns=['转向半径'])
        # logging(angle_df.head())
        # logging(angle_df.info())
        # logging(radius_df.head())
        # logging(radius_df.info())
        
        output_df = pd.concat([self.df.iloc[:self.num, :], angle_df, radius_df], axis=1)
        # logging(output_df.info())
        logging('\n数据行数和列数:\n')
        logging(output_df.shape)
        logging('\n有效角度数据数量:\n')
        logging(angle_df['转向角度'].count())
        logging('\n有效转向半径数量:\n')
        logging(radius_df['转向半径'].count())
        # 打开Excel文件
        book = pyxl.load_workbook(self.filename_out)

        # 新建工作表并将DataFrame对象写入其中
        writer = pd.ExcelWriter(self.filename_out, engine='openpyxl')
        writer.book = book
        output_df.to_excel(writer, sheet_name=self.sheetname)

        # 保存Excel文件
        writer.save()
        self.done+=1
        
        
            
            
            